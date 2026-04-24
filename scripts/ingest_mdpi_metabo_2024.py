from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import shutil
import zipfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from scipy import stats


ROOT = Path(__file__).resolve().parents[1]
SOURCE_ID = "MDPI_METABO_2024_LT_GRAFT_PATHOLOGY"
RAW_DIR = ROOT / "data" / "raw" / SOURCE_ID
UNPACKED_DIR = RAW_DIR / "unpacked"
PROCESSED_DIR = ROOT / "data" / "processed" / SOURCE_ID

ZIP_URL = "https://mdpi-res.com/d_attachment/metabolites/metabolites-14-00254/article_deploy/metabolites-14-00254-s001.zip"
PDF_URL = "https://mdpi-res.com/d_attachment/metabolites/metabolites-14-00254/article_deploy/metabolites-14-00254.pdf"
ZIP_PATH = RAW_DIR / "metabolites-14-00254-s001.zip"
PDF_PATH = RAW_DIR / "metabolites-14-00254.pdf"
TABLE_PATH = UNPACKED_DIR / "Table S4.xlsx"
LEGACY_RAW_DIR = ROOT / "data" / "raw" / "MDPI_Metabo_2024"

CLASS_MAP = {
    "MASH": "post_transplant_mash",
    "TCMR": "TCMR",
    "Biliary": "biliary_complication",
}

PAIRWISE_CONTRASTS = [
    ("TCMR", "biliary_complication"),
    ("TCMR", "post_transplant_mash"),
    ("biliary_complication", "post_transplant_mash"),
]


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def download_if_missing(url: str, path: Path) -> None:
    if path.exists():
        return
    ensure_dir(path.parent)
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request) as response, path.open("wb") as handle:
        shutil.copyfileobj(response, handle)


def hydrate_from_legacy_cache() -> None:
    if not LEGACY_RAW_DIR.exists():
        return
    for source, target in [
        (LEGACY_RAW_DIR / "metabolites-14-00254-s001.zip", ZIP_PATH),
        (LEGACY_RAW_DIR / "metabolites-14-00254.pdf", PDF_PATH),
        (LEGACY_RAW_DIR / "unpacked" / "Table S4.xlsx", TABLE_PATH),
    ]:
        if source.exists() and not target.exists():
            ensure_dir(target.parent)
            shutil.copy2(source, target)


def unpack_if_needed(zip_path: Path, target_dir: Path) -> None:
    ensure_dir(target_dir)
    if TABLE_PATH.exists():
        return
    with zipfile.ZipFile(zip_path) as archive:
        archive.extractall(target_dir)


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return slug or "feature"


def round_or_none(value: float | None, digits: int = 6) -> float | None:
    if value is None or not math.isfinite(value):
        return None
    return round(value, digits)


def parse_numeric(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.upper() in {"NA", "N/A"} or text == "< LOD":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def numeric_summary(values: list[float]) -> dict[str, Any]:
    if not values:
        return {"n": 0}
    ordered = sorted(values)
    mid = len(ordered) // 2
    median = ordered[mid] if len(ordered) % 2 else (ordered[mid - 1] + ordered[mid]) / 2
    mean = sum(values) / len(values)
    variance = sum((x - mean) ** 2 for x in values) / max(len(values) - 1, 1)
    return {
        "n": len(values),
        "mean": round_or_none(mean),
        "median": round_or_none(median),
        "min": round_or_none(min(values)),
        "max": round_or_none(max(values)),
        "sd": round_or_none(math.sqrt(variance)),
    }


def benjamini_hochberg(rows: list[dict[str, Any]], p_key: str = "p_value") -> None:
    indexed = [(idx, row.get(p_key)) for idx, row in enumerate(rows) if row.get(p_key) is not None]
    indexed.sort(key=lambda item: item[1])
    total = len(indexed)
    adjusted = [None] * len(rows)
    running = 1.0
    for rank, (idx, p_value) in reversed(list(enumerate(indexed, start=1))):
        candidate = min(running, p_value * total / rank)
        running = candidate
        adjusted[idx] = round_or_none(candidate)
    for idx, row in enumerate(rows):
        row["adj_p_value_bh"] = adjusted[idx]


def build_records() -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, Any], dict[str, Any], dict[str, Any]]:
    workbook = load_workbook(TABLE_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    sample_ids = [str(value) for value in rows[0][1:] if value is not None]
    class_labels = [CLASS_MAP[str(value)] for value in rows[1][1:] if value is not None]

    samples = []
    for sample_id, clinical_state in zip(sample_ids, class_labels, strict=True):
        samples.append(
            {
                "sample_accession": sample_id,
                "title": f"Serum metabolomics {sample_id}",
                "sample_origin": "recipient_serum",
                "clinical_state": clinical_state,
                "assay_modality": "metabolomics",
                "source_table": "Table S4.xlsx",
            }
        )

    state_counts = dict(sorted(Counter(class_labels).items()))
    sample_summary = {
        "study_accession": SOURCE_ID,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "sample_count": len(samples),
        "by_clinical_state": state_counts,
        "by_sample_origin": {"recipient_serum": len(samples)},
        "by_assay_modality": {"metabolomics": len(samples)},
    }

    cohort_summary = {
        "source_id": SOURCE_ID,
        "generated_at_utc": sample_summary["generated_at_utc"],
        "patient_count": len(samples),
        "sample_count": len(samples),
        "clinical_group_counts": state_counts,
        "pathology_groups": [
            {"clinical_state": state, "sample_count": count}
            for state, count in state_counts.items()
        ],
        "measurement_unit": "uM",
        "sample_origin": "recipient_serum",
    }

    feature_rows: list[dict[str, Any]] = []
    contrast_tables = {f"{case}_vs_{control}": [] for case, control in PAIRWISE_CONTRASTS}

    for metabolite_row in rows[2:]:
        metabolite = str(metabolite_row[0]).strip()
        raw_values = [parse_numeric(value) for value in metabolite_row[1:]]
        values = [value for value in raw_values if value is not None]
        grouped: dict[str, list[float]] = {state: [] for state in sorted(set(class_labels))}
        per_sample = []
        for sample_id, clinical_state, value in zip(sample_ids, class_labels, raw_values, strict=True):
            if value is not None:
                grouped[clinical_state].append(value)
            per_sample.append(
                {
                    "sample_accession": sample_id,
                    "clinical_state": clinical_state,
                    "value": round_or_none(value),
                }
            )

        feature = {
            "feature_id": slugify(metabolite),
            "feature_type": "metabolite",
            "display_name": metabolite,
            "modality": "metabolomics",
            "source_id": SOURCE_ID,
            "source_table": "Table S4.xlsx",
            "assay_scope": "absolute_concentration_uM",
            "measurement": "uM",
            "patient_count": len(samples),
            "sample_count": len(values),
            "all_samples": numeric_summary(values),
            "sample_values": per_sample,
            "contrasts": {},
            "limitations": [
                "Absolute metabolite concentrations were parsed from a public supplementary spreadsheet.",
                "This is a single-center serum cohort spanning TCMR, biliary complications, and post-transplant MASH.",
                "Contrasts are exploratory Welch t-tests and do not reproduce the paper's full machine-learning classifier.",
            ],
        }

        for state, state_values in grouped.items():
            feature[state] = numeric_summary(state_values)

        for case_state, control_state in PAIRWISE_CONTRASTS:
            case_values = grouped[case_state]
            control_values = grouped[control_state]
            statistic, p_value = stats.ttest_ind(case_values, control_values, equal_var=False)
            contrast_id = f"{case_state}_vs_{control_state}"
            contrast_row = {
                "feature_id": feature["feature_id"],
                "display_name": metabolite,
                "case_state": case_state,
                "control_state": control_state,
                "case_n": len(case_values),
                "control_n": len(control_values),
                "case_mean": round_or_none(sum(case_values) / len(case_values)),
                "control_mean": round_or_none(sum(control_values) / len(control_values)),
                "mean_difference": round_or_none((sum(case_values) / len(case_values)) - (sum(control_values) / len(control_values))),
                "effect_scale": "absolute_concentration_uM",
                "statistic_welch_t": round_or_none(statistic),
                "p_value": round_or_none(p_value),
            }
            contrast_tables[contrast_id].append(contrast_row)
            feature["contrasts"][contrast_id] = contrast_row.copy()

        feature_rows.append(feature)

    for contrast_id, rows_for_contrast in contrast_tables.items():
        benjamini_hochberg(rows_for_contrast)
        rows_by_feature = {row["feature_id"]: row for row in rows_for_contrast}
        for feature in feature_rows:
            feature["contrasts"][contrast_id]["adj_p_value_bh"] = rows_by_feature[feature["feature_id"]]["adj_p_value_bh"]

    variances = []
    for feature in feature_rows:
        mean = feature["all_samples"]["mean"]
        values = [item["value"] for item in feature["sample_values"] if item["value"] is not None]
        variance = sum((value - mean) ** 2 for value in values) / max(len(values) - 1, 1)
        variances.append((feature["display_name"], variance))
    variances.sort(key=lambda item: item[1], reverse=True)

    metabolomics_summary = {
        "source_id": SOURCE_ID,
        "generated_at_utc": sample_summary["generated_at_utc"],
        "modality": "metabolomics",
        "feature_count": len(feature_rows),
        "sample_count": len(samples),
        "measurement_unit": "uM",
        "clinical_group_counts": state_counts,
        "pairwise_contrasts": list(contrast_tables),
        "top_features_by_variance": [
            {"display_name": name, "variance": round_or_none(variance)}
            for name, variance in variances[:25]
        ],
    }

    metabolomics_features = {
        "source_id": SOURCE_ID,
        "generated_at_utc": sample_summary["generated_at_utc"],
        "modality": "metabolomics",
        "feature_count": len(feature_rows),
        "absolute_concentration_uM": {
            "feature_count": len(feature_rows),
            "features": feature_rows,
        },
    }

    provenance = {
        "source_id": SOURCE_ID,
        "generated_at_utc": sample_summary["generated_at_utc"],
        "parser": "scripts/ingest_mdpi_metabo_2024.py",
        "source_urls": {
            "article_pdf": PDF_URL,
            "supplementary_zip": ZIP_URL,
        },
        "source_files": [
            {
                "path": str(PDF_PATH.relative_to(ROOT)),
                "bytes": PDF_PATH.stat().st_size,
                "sha256": sha256_file(PDF_PATH),
            },
            {
                "path": str(ZIP_PATH.relative_to(ROOT)),
                "bytes": ZIP_PATH.stat().st_size,
                "sha256": sha256_file(ZIP_PATH),
            },
            {
                "path": str(TABLE_PATH.relative_to(ROOT)),
                "bytes": TABLE_PATH.stat().st_size,
                "sha256": sha256_file(TABLE_PATH),
            },
        ],
        "methods": [
            "Downloaded the public article PDF and supplementary ZIP from mdpi-res.com.",
            "Parsed Supplementary Table S4.xlsx absolute metabolite concentrations for 55 serum samples.",
            "Standardized pathology groups to TCMR, biliary_complication, and post_transplant_mash.",
            "Computed exploratory pairwise Welch t-tests with Benjamini-Hochberg correction across metabolites for each contrast.",
        ],
        "limitations": [
            "This is a single-center post-transplant serum metabolomics cohort with modest sample size.",
            "The public supplementary spreadsheet supports feature-level metabolite evidence, but not full reproduction of the paper's machine-learning workflow.",
            "Deidentified clinical covariates beyond pathology class are not public in the supplementary spreadsheet.",
        ],
        "outputs": [
            "data/processed/MDPI_METABO_2024_LT_GRAFT_PATHOLOGY/samples.json",
            "data/processed/MDPI_METABO_2024_LT_GRAFT_PATHOLOGY/sample_summary.json",
            "data/processed/MDPI_METABO_2024_LT_GRAFT_PATHOLOGY/cohort_summary.json",
            "data/processed/MDPI_METABO_2024_LT_GRAFT_PATHOLOGY/metabolomics_summary.json",
            "data/processed/MDPI_METABO_2024_LT_GRAFT_PATHOLOGY/metabolomics_features.json",
            "data/processed/MDPI_METABO_2024_LT_GRAFT_PATHOLOGY/source_file_inventory.json",
            "data/processed/MDPI_METABO_2024_LT_GRAFT_PATHOLOGY/provenance.json",
        ],
    }

    inventory = {
        "source_id": SOURCE_ID,
        "generated_at_utc": sample_summary["generated_at_utc"],
        "raw_files": provenance["source_files"],
        "supplementary_members": [item.filename for item in zipfile.ZipFile(ZIP_PATH).infolist()],
    }

    return samples, sample_summary, cohort_summary, metabolomics_summary, metabolomics_features, provenance, inventory


def write_json(path: Path, payload: Any) -> None:
    ensure_dir(path.parent)
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest MDPI 2024 liver graft pathology serum metabolomics supplementary data.")
    parser.add_argument("--force-download", action="store_true")
    args = parser.parse_args()

    ensure_dir(RAW_DIR)
    ensure_dir(PROCESSED_DIR)
    hydrate_from_legacy_cache()
    if args.force_download:
        if ZIP_PATH.exists():
            ZIP_PATH.unlink()
        if PDF_PATH.exists():
            PDF_PATH.unlink()
        if TABLE_PATH.exists():
            TABLE_PATH.unlink()

    download_if_missing(ZIP_URL, ZIP_PATH)
    download_if_missing(PDF_URL, PDF_PATH)
    unpack_if_needed(ZIP_PATH, UNPACKED_DIR)

    samples, sample_summary, cohort_summary, metabolomics_summary, metabolomics_features, provenance, inventory = build_records()
    write_json(PROCESSED_DIR / "samples.json", samples)
    write_json(PROCESSED_DIR / "sample_summary.json", sample_summary)
    write_json(PROCESSED_DIR / "cohort_summary.json", cohort_summary)
    write_json(PROCESSED_DIR / "metabolomics_summary.json", metabolomics_summary)
    write_json(PROCESSED_DIR / "metabolomics_features.json", metabolomics_features)
    write_json(PROCESSED_DIR / "provenance.json", provenance)
    write_json(PROCESSED_DIR / "source_file_inventory.json", inventory)
    print(
        json.dumps(
            {
                "source_id": SOURCE_ID,
                "sample_count": sample_summary["sample_count"],
                "feature_count": metabolomics_summary["feature_count"],
                "clinical_group_counts": cohort_summary["clinical_group_counts"],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
